"""
4M IT Equipment Reports
"""
import streamlit as st
import pandas as pd
import openpyxl
from io import BytesIO
from datetime import datetime
from database_it import (
    IT_EQUIPMENT_TYPES,
    get_rdcs, get_it_shipments,
    get_rdc_requirements_dict, get_shipped_totals_per_rdc,
    get_it_stock_dict, get_total_requirements, get_all_shipment_items,
    get_upcoming_rdc_launches
)
from database import get_threshold
from components import render_stat_card, render_section_title
from email_drafts import generate_it_mailto


def build_rdc_report(rdcs_df):
    """Build per-RDC report with all equipment"""
    rows = []
    for _, rdc in rdcs_df.iterrows():
        requirements = get_rdc_requirements_dict(rdc['id'])
        shipped = get_shipped_totals_per_rdc(rdc['id'])

        launch = rdc.get('launch_date')
        launch_str = str(launch) if launch else '-'
        transport = '✅' if rdc.get('transportation_ready') else '❌'

        row = {
            'RDC': rdc['name'],
            'Location': rdc['location'] or '-',
            'Launch Date': launch_str,
            'Transport Ready': transport,
        }

        total_req = 0
        total_ship = 0
        for item in IT_EQUIPMENT_TYPES:
            req = int(requirements.get(item, 0))
            ship = int(shipped.get(item, 0))
            pending = max(0, req - ship)
            short = item.replace('Wireless Scanner - ', 'WS-').replace('Charger Wireless Scanner - ', 'Charger-').replace('Yubikey - Security Key', 'Yubikey').replace('Zebra ZD621', 'Zebra')
            row[f'{short} Req'] = req
            row[f'{short} Ship'] = ship
            row[f'{short} Pend'] = pending
            total_req += req
            total_ship += ship

        row['Total Req'] = total_req
        row['Total Ship'] = total_ship
        row['Total Pending'] = max(0, total_req - total_ship)

        rows.append(row)

    return pd.DataFrame(rows)


def build_equipment_summary():
    """Build per-equipment type summary"""
    stocks = get_it_stock_dict()
    requirements = get_total_requirements()
    all_items = get_all_shipment_items()

    rows = []
    for item in IT_EQUIPMENT_TYPES:
        stock = stocks.get(item, 0)
        req = requirements.get(item, 0)
        shipped = int(all_items[all_items['equipment_type'] == item]['quantity'].sum()) if not all_items.empty else 0
        pending = max(0, req - shipped)

        rows.append({
            'Equipment': item,
            'Current Stock': stock,
            'Total Required': req,
            'Total Shipped': shipped,
            'Pending': pending,
            'Stock Status': '✅ OK' if stock >= pending else ('⚠️ Low' if stock > 0 else '❌ Out')
        })

    return pd.DataFrame(rows)


def build_rdc_progress_list(rdcs_df):
    """Build list of RDCs with progress info for email"""
    progress = []
    for _, rdc in rdcs_df.iterrows():
        requirements = get_rdc_requirements_dict(rdc['id'])
        shipped = get_shipped_totals_per_rdc(rdc['id'])

        total_req = sum(requirements.values())
        total_ship = sum(shipped.values())
        total_pending = max(0, total_req - total_ship)
        pct = (total_ship / total_req * 100) if total_req > 0 else 0

        progress.append({
            'name': rdc['name'],
            'location': rdc['location'] or 'N/A',
            'total_req': total_req,
            'total_ship': total_ship,
            'total_pending': total_pending,
            'pct': pct,
        })
    return progress


def build_shipped_totals():
    """Get total shipped per equipment type"""
    all_items = get_all_shipment_items()
    totals = {t: 0 for t in IT_EQUIPMENT_TYPES}
    if not all_items.empty:
        for t in IT_EQUIPMENT_TYPES:
            totals[t] = int(all_items[all_items['equipment_type'] == t]['quantity'].sum())
    return totals


def export_to_excel(rdc_df, equipment_df, shipments_df):
    """Export all data to multi-sheet Excel"""
    output = BytesIO()
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "RDCs Report"
    ws1.append(list(rdc_df.columns))
    for _, row in rdc_df.iterrows():
        ws1.append(list(row))

    ws2 = wb.create_sheet("Equipment Summary")
    ws2.append(list(equipment_df.columns))
    for _, row in equipment_df.iterrows():
        ws2.append(list(row))

    if not shipments_df.empty:
        ws3 = wb.create_sheet("Shipments")
        ship_cols = ['id', 'rdc_name', 'date', 'scheduled_date', 'delivery_status',
                     'receiver_name', 'receiver_contact', 'notes']
        available = [c for c in ship_cols if c in shipments_df.columns]
        ws3.append(available)
        for _, row in shipments_df.iterrows():
            ws3.append([row.get(c, '') for c in available])

    wb.save(output)
    output.seek(0)
    return output


def style_pending(val):
    """Style pending cells"""
    try:
        v = int(val)
        if v > 0:
            return 'color: #e74c3c; font-weight: 700;'
        elif v == 0:
            return 'color: #27ae60; font-weight: 600;'
    except (ValueError, TypeError):
        pass
    return ''


def render():
    """Render the IT Reports page"""
    st.markdown("# 📈 4M IT Equipment Reports")
    st.caption("Detailed reports on RDCs and equipment shipments")

    rdcs_df = get_rdcs()
    shipments_df = get_it_shipments()

    if rdcs_df.empty:
        st.info("📭 No RDCs to report.")
        return

    rdc_report = build_rdc_report(rdcs_df)
    equipment_summary = build_equipment_summary()

    # Summary cards
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        render_stat_card("Total RDCs", len(rdcs_df), "card-stores", "bi-building")
    with c2:
        render_stat_card("Total Required", int(rdc_report['Total Req'].sum()), "card-30d", "bi-clipboard-check")
    with c3:
        render_stat_card("Total Shipped", int(rdc_report['Total Ship'].sum()), "card-shipments", "bi-truck")
    with c4:
        render_stat_card("Total Pending", int(rdc_report['Total Pending'].sum()), "card-40d", "bi-hourglass-split")

    # === EMAIL DRAFT SECTION ===
    render_section_title("📧 Email Daily Report")

    st.markdown("""
    <div style="background: rgba(52, 152, 219, 0.1); padding: 14px 18px; border-radius: 10px; 
                border-left: 4px solid #3498db; margin-bottom: 16px;">
        💡 Click the button below to open your default email app with a pre-filled IT equipment daily report.
        <br>You can review and edit before sending.
    </div>
    """, unsafe_allow_html=True)

    # Gather data for email
    stocks = get_it_stock_dict()
    threshold = get_threshold()
    requirements = get_total_requirements()
    shipped_totals = build_shipped_totals()
    upcoming_launches = get_upcoming_rdc_launches(days_ahead=4)
    rdc_progress = build_rdc_progress_list(rdcs_df)

    mailto_link = generate_it_mailto(
        stocks, threshold, requirements, shipped_totals,
        rdcs_df, shipments_df, upcoming_launches, rdc_progress
    )

    c1, c2, c3 = st.columns([1, 1, 1])
    with c2:
        st.markdown(
            f'''
            <a href="{mailto_link}" style="
                display: block;
                text-align: center;
                background: linear-gradient(135deg, #9b59b6 0%, #6c3483 100%);
                color: white !important;
                padding: 12px 24px;
                border-radius: 10px;
                text-decoration: none;
                font-weight: 700;
                box-shadow: 0 4px 14px rgba(155, 89, 182, 0.35);
                transition: all 0.3s ease;
                margin: 10px 0;
            " onmouseover="this.style.transform='translateY(-3px)'; this.style.boxShadow='0 6px 20px rgba(155, 89, 182, 0.5)';"
            onmouseout="this.style.transform='translateY(0)'; this.style.boxShadow='0 4px 14px rgba(155, 89, 182, 0.35)';">
                📧 Generate IT Equipment Report Email
            </a>
            ''',
            unsafe_allow_html=True
        )

    # Equipment Summary
    render_section_title("📦 Equipment Summary")

    pending_cols = ['Pending']
    available_pending = [c for c in pending_cols if c in equipment_summary.columns]
    styled_eq = equipment_summary.style.map(style_pending, subset=available_pending)
    st.dataframe(styled_eq, use_container_width=True, hide_index=True)

    # RDC Detailed Report
    render_section_title("🏢 Per-RDC Detailed Report")

    pending_cols_rdc = [c for c in rdc_report.columns if 'Pend' in c]
    styled_rdc = rdc_report.style.map(style_pending, subset=pending_cols_rdc)
    st.dataframe(styled_rdc, use_container_width=True, hide_index=True)

    # Shipments summary
    if not shipments_df.empty:
        render_section_title("🚚 Recent Shipments")
        display_cols = ['id', 'rdc_name', 'date', 'scheduled_date', 'delivery_status',
                        'receiver_name', 'receiver_contact', 'notes']
        available = [c for c in display_cols if c in shipments_df.columns]
        display_df = shipments_df[available].copy()
        rename_map = {
            'id': 'ID',
            'rdc_name': 'RDC',
            'date': 'Ship Date',
            'scheduled_date': 'Scheduled',
            'delivery_status': 'Status',
            'receiver_name': 'Receiver',
            'receiver_contact': 'Contact',
            'notes': 'Notes'
        }
        display_df = display_df.rename(columns=rename_map)
        st.dataframe(display_df, use_container_width=True, hide_index=True)

    # Excel download
    render_section_title("📥 Export")
    output = export_to_excel(rdc_report, equipment_summary, shipments_df)

    c1, c2, c3 = st.columns([1, 1, 1])
    with c2:
        st.download_button(
            label="📥 Download Full Excel Report",
            data=output,
            file_name=f"it_equipment_report_{datetime.now().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
